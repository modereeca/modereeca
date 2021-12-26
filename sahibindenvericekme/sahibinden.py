from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
siralama = "sorting_date_desc"
browser = webdriver.Chrome()
kiralik_urls = ["https://www.sahibinden.com/kiralik-daire/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-residence/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-mustakil-ev/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-villa/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-ciftlik-evi/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-yali-dairesi/izmir/sahibinden",
"https://www.sahibinden.com/kiralik-yazlik/izmir/sahibinden" 
"https://www.sahibinden.com/kiralik-daire/izmir/insaat-firmasindan"
"https://www.sahibinden.com/kiralik-residence/izmir/insaat-firmasindan"
"https://www.sahibinden.com/kiralik-mustakil-ev/izmir/insaat-firmasindan"
"https://www.sahibinden.com/kiralik-villa/izmir/insaat-firmasindan"
"https://www.sahibinden.com/kiralik-yazlik/izmir/insaat-firmasindan"
]

basliklar = ("İSİM", "İLAN TARİHİ", "EMLAK TİPİ", "M2 (BRÜT)", "M2 (NET)", "ODA SAYISI", "BİNA YAŞI", "BULUNDUĞU KAT", "KAT SAYISI", "ISITMA", "FİYAT", "İLÇE", "MAHALLE", "KİMDEN", "TELEFON NUMARASI 1", "İLAN LİNKİ")
a = 1

for baslik in basliklar:
    ws.cell(row=1, column=a).value = baslik
    a += 1


# SAHİBİNDEN KİRALIK ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#   daire = https://www.sahibinden.com/kiralik-daire/izmir/sahibinden
#   residence = https://www.sahibinden.com/kiralik-residence/izmir/sahibinden
#   müstakil ev = https://www.sahibinden.com/kiralik-mustakil-ev/izmir/sahibinden
#   villa = https://www.sahibinden.com/kiralik-villa/izmir/sahibinden
#   çiftlik evi = https://www.sahibinden.com/kiralik-ciftlik-evi/izmir/sahibinden
#   yalı dairesi = https://www.sahibinden.com/kiralik-yali-dairesi/izmir/sahibinden
#   yazlık = https://www.sahibinden.com/kiralik-yazlik/izmir/sahibinden

# SAHİBİNDEN SATILIK ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#   daire = https://www.sahibinden.com/satilik-daire/izmir/sahibinden?sorting=date_desc
#   residence = https://www.sahibinden.com/satilik-residence/izmir/sahibinden?sorting=date_desc
#   müstakil ev = https://www.sahibinden.com/satilik-mustakil-ev/izmir/sahibinden?sorting=date_desc
#   villa = https://www.sahibinden.com/satilik-villa/izmir/sahibinden?sorting=date_desc
#   çiftlik evi = https://www.sahibinden.com/satilik-ciftlik-evi/izmir?sorting=date_desc&a27=38460
#   köşk & konak = https://www.sahibinden.com/satilik-kosk-konak/izmir?sorting=date_desc&a27=38460
#   yalı dairesi = https://www.sahibinden.com/satilik-yali-dairesi/izmir?sorting=date_desc&a27=38460
#   yazlık = https://www.sahibinden.com/satilik-yazlik/izmir/sahibinden?sorting=date_desc
#   prefabrik ev = https://www.sahibinden.com/satilik-prefabrik-ev/izmir?sorting=date_desc&a27=38460
#   kooperatif = https://www.sahibinden.com/satilik-kooperatif/izmir/sahibinden?sorting=date_desc

# İNŞAAT FİRMASINDAN KİRALIK ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#   daire = https://www.sahibinden.com/kiralik-daire/izmir/insaat-firmasindan
#   residence = https://www.sahibinden.com/kiralik-residence/izmir/insaat-firmasindan
#   müstakil ev = https://www.sahibinden.com/kiralik-mustakil-ev/izmir/insaat-firmasindan
#   villa = https://www.sahibinden.com/kiralik-villa/izmir/insaat-firmasindan
#   yazlık = https://www.sahibinden.com/kiralik-yazlik/izmir/insaat-firmasindan

# İNŞAAT FİRMASINDAN SATILIK ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#   daire = https://www.sahibinden.com/satilik-daire/izmir/insaat-firmasindan?sorting=date_desc
#   residence = https://www.sahibinden.com/satilik-residence/izmir/insaat-firmasindan?sorting=date_desc
#   müstakil ev = https://www.sahibinden.com/satilik-mustakil-ev/izmir/insaat-firmasindan?sorting=date_desc
#   villa = https://www.sahibinden.com/satilik-villa/izmir/insaat-firmasindan?sorting=date_desc
#   çiftlik evi = https://www.sahibinden.com/satilik-ciftlik-evi/izmir/insaat-firmasindan?sorting=date_desc
#   yalı dairesi = https://www.sahibinden.com/satilik-yali-dairesi/izmir/insaat-firmasindan?a83486=654966&sorting=date_desc
#   yazlık = https://www.sahibinden.com/satilik-yazlik/izmir/insaat-firmasindan?a83486=654966&sorting=date_desc&

query = "?pagingOffset="
r = 2
for url in kiralik_urls:
    Offset = 0
    while True:
        browser.get(url + query + str(Offset))
        time.sleep(1)
        if len(browser.find_elements_by_class_name("classifiedTitle")) == 0:
            break
        else:
            for i in range(len(browser.find_elements_by_class_name("classifiedTitle"))):
                eleman = browser.find_elements_by_class_name("classifiedTitle")[i]
                eleman.send_keys(Keys.ENTER)
                time.sleep(1)
                c = 1
                isim = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[3]/div[1]/div/div[1]/h5')
                ws.cell(column=c, row=r).value = isim.text
                c += 1
                ilantarihi = browser.find_element_by_xpath("//*[@id='classifiedDetail']/div/div[2]/div[2]/ul/li[2]/span")
                ws.cell(column=c, row=r).value = ilantarihi.text
                c += 1
                emlaktipi = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[3]/span')
                ws.cell(column=c, row=r).value = emlaktipi.text
                c += 1
                metrekarebrut = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[4]/span')
                ws.cell(column=c, row=r).value = metrekarebrut.text
                c += 1
                metrekarenet = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[5]/span')
                ws.cell(column=c, row=r).value = metrekarenet.text
                c += 1
                odasayisi = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[6]/span')
                ws.cell(column=c, row=r).value = odasayisi.text
                c += 1
                binayasi = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[7]/span')
                ws.cell(column=c, row=r).value = binayasi.text
                c += 1
                bulundugukat = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[8]/span')
                ws.cell(column=c, row=r).value = bulundugukat.text
                c += 1
                katsayisi = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[9]/span')
                ws.cell(column=c, row=r).value = katsayisi.text
                c += 1
                isitma = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[10]/span')
                ws.cell(column=c, row=r).value = isitma.text
                c += 1
                fiyat = browser.find_element_by_xpath("//*[@id='classifiedDetail']/div/div[2]/div[2]/h3")
                ws.cell(column=c, row=r).value = fiyat.text
                c += 1
                ilce = browser.find_element_by_xpath("//*[@id='classifiedDetail']/div/div[2]/div[2]/h2/a[2]")
                ws.cell(column=c, row=r).value = ilce.text
                c += 1
                mahalle = browser.find_element_by_xpath("//*[@id='classifiedDetail']/div/div[2]/div[2]/h2/a[3]")
                ws.cell(column=c, row=r).value = mahalle.text
                c += 1
                if len(browser.find_elements_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[19]/span')) > 0:
                    kimden = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[19]/span')
                    ws.cell(column=c, row=r).value = kimden.text
                    c += 1
                else:
                    kimden = browser.find_element_by_xpath('//*[@id="classifiedDetail"]/div/div[2]/div[2]/ul/li[17]/span')
                    ws.cell(column=c, row=r).value = kimden.text
                    c += 1
                if browser.find_elements_by_xpath("//*[@id='phoneInfoPart']/li"):
                    if len(browser.find_elements_by_xpath("//*[@id='phoneInfoPart']/li")) == 1:
                        numara = browser.find_element_by_xpath("//*[@id='phoneInfoPart']/li")
                        ws.cell(column=c, row=r).value = numara.text[3::]
                    elif len(browser.find_elements_by_xpath("//*[@id='phoneInfoPart']/li")) == 2:
                        numara = browser.find_element_by_xpath("//*[@id='phoneInfoPart']/li[2]")
                        ws.cell(column=c, row=r).value = numara.text[3::]
                    elif len(browser.find_elements_by_xpath("//*[@id='phoneInfoPart']/li")) == 3:
                        numara = browser.find_element_by_xpath("//*[@id='phoneInfoPart']/li[3]")
                        ws.cell(column=c, row=r).value = numara.text[3::]
                elif browser.find_elements_by_id("showPhoneNumberLink"):
                    pass
                elif browser.find_elements_by_class_name("userRegistrationDate"):
                    pass
                c += 1
                ilanlinki = browser.current_url
                ws.cell(column=c, row=r).value = ilanlinki
                r += 1
                browser.get(url + query + str(Offset))

            Offset += 20
            time.sleep(1)
wb.save("kiralik_ilanlar.xlsx")
browser.quit()
